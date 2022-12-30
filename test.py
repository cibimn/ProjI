import openpyxl as o
import json
import uuid
import datetime
import os
import pandas as pd

def create():
    excel = "C:\\Users\\cibim\\Downloads\\brandclassids.xlsx"
    wb = o.load_workbook(excel)
    ws = wb.active
    mr = ws.max_row
    a = []
    d = dict({})
    for k in range(2,mr+1):
        d[ws.cell(row=k, column=1).value] = str(ws.cell(row=k, column=2).value)
        # d['RowKey'] = str(uuid.uuid4())
        # d['ITEM_BRAND_NAME_ID'] = ws.cell(row=k, column=1).value
        # d['ITEM_CATEGORY_NAME'] = ws.cell(row=k, column=3).value
        # d['ITEM_OPERATIONAL_NAME'] = ws.cell(row=k, column=2).value
        # d['COUNTRY'] = ws.cell(row=k, column=1).value
        # d['ITEM_BRAND_NAME_ID'] = ws.cell(row=k, column=2).value
        # d['ITEM_CATEGORY_NAME'] = ws.cell(row=k, column=3).value
        # d['ITEM_OPERATIONAL_NAME'] = ws.cell(row=k, column=4).value
        # d['SF_GLOBAL_BRAND'] = ws.cell(row=k, column=5).value
        # d['LEN'] = ws.cell(row=k, column=6).value
        a.append(d)
    print(len(a))
    #D:\Projects\Singularium\droofunctions_app\drooapis\roodetect\brand_predict
    filename = "D:\\Projects\\Singularium\\droofunctions_app\\drooapis\\roodetect\\brand_predict\\"+"brand_config"+".json"
    json_object = json.dumps(d, indent = 4)    
    with open(filename, "w") as outfile:
        outfile.write(json_object)
    # json_object = json.dumps(a, indent = 4)
    # print(json_object)
    # data_list = [a[x:x+10000] for x in range(0, len(a), 10000)]
    # for data in data_list:
    #     json_object = json.dumps(data, indent = 2)
    #     d = datetime.datetime.now()
    #     file_name = str(d.timestamp())+".json"
    #     with open(file_name, "w") as outfile:
    #         outfile.write(json_object)
    # data_list = [a[x:x+27693] for x in range(0, len(a), 27693)]
    # ae = 0
    # for data in data_list:
    #     print(len(data))
    #     ae+=len(data)
        # filename = "D:\\JSON\\sample0"+str(ae)+".json"
        # json_object = json.dumps(data, indent = 4)    
        # with open(filename, "w") as outfile:
        #     outfile.write(json_object)
        # ae = ae+1
    # print(ae)

def onefile():
    directory = "D:\\Output\\New\\"
    a = []
    for x in os.listdir(directory):
        if x.endswith(".json"):
            # Prints only text file present in My Folder
            a.append(x)
    de = []
    for r in range(0, len(a)):
        print(r)
        df = pd.read_json(directory+a[r])
        de.append(df)
    df2 = pd.concat(de)
    df2.to_csv("D:\\TransNe.csv",index=False)
create()



def rangecalc(i):
    a = [k for k in range(1, 20000)]
    b = [a[e:min(e+10, len(a))] for e in range(0, len(a), 10)]
    for ke in b:
        if i in ke:
            val = (b.index(ke) +1)*10
            print(val)
            break
    return val
    
rangecalc(191)